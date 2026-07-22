"""
Erstellt pro Klasse eine Planungs-Excel-Datei aus den Stundenplan-Exporten in
output_files/ (siehe pdf_timetable_to_xlsx.py).

Pro Klasse entsteht genau EINE Excel-Datei. Darin gibt es entweder
  - EIN Arbeitsblatt fürs ganze Jahr, wenn sich der Stundenplan der Klasse
    (Fächer, Wochentage, Lektionenzahl) über alle Phasen hinweg nicht
    ändert, oder
  - EIN Arbeitsblatt PRO PHASE, wenn er sich ändert (inkl. Phasen, in denen
    die Klasse gar nicht unterrichtet wird).

InL- und Konf-Lektionen aus den Stundenplan-Exporten werden dabei ignoriert.
Bei Klassen mit mehreren Phasen-Arbeitsblättern werden stattdessen pro
Kalenderwoche 2 fixe InL-Lektionen an den Wochenanfang gesetzt (statt eines
Datums steht dort "InL").

Schulkalender (Ferien, unterrichtsfreie Tage, Spezialtage, Spezialwochen)
und Gestaltungswerte (Schrift, Farben, Masse) stehen in config.py, da sie
nicht aus den Stundenplan-Exporten hervorgehen und sich jedes Schuljahr
ändern können.

Diese Datei orchestriert nur noch: Stundenplan-Exporte einlesen
(stundenplan_parsing.py), pro Klasse prüfen ob der Stundenplan übers Jahr
konstant bleibt, und die Excel-Arbeitsblätter befüllen (rendering.py).
"""

import re
from pathlib import Path

import openpyxl

from config import BASIS_JAHR, FERIEN, UNTERRICHTSFREIETAGE, SPEZIALTAGE, SPEZIALWOCHEN
from rendering import fuelle_arbeitsblatt
from stundenplan_parsing import lade_phasen


def klassenstufe_von_klasse(klasse):
    """Führende Ziffern eines Klassencodes, z. B. '3ACPQ' -> 3."""
    m = re.match(r"^(\d+)", klasse)
    return int(m.group(1)) if m else None


def klassen_signatur(phase, klasse, faecher_namen):
    """(fach, sortierte (wochentag, perioden)-Paare) für jedes Fach der Klasse in dieser Phase."""
    return tuple(
        (fach, tuple(sorted(phase["lektionen"].get(klasse, {}).get(fach, {}).items())))
        for fach in faecher_namen
    )


def ist_konstant(phasen_teil, klasse, faecher_namen):
    """True, wenn die Klasse in JEDER Phase von phasen_teil unterrichtet wird
    und dabei überall denselben Stundenplan (Fächer, Wochentage, Lektionenzahl)
    hat."""
    signaturen = [klassen_signatur(p, klasse, faecher_namen) for p in phasen_teil]
    hat_lektionen = [any(eintraege for _, eintraege in sig) for sig in signaturen]
    return all(hat_lektionen) and all(sig == signaturen[0] for sig in signaturen)


def faecher_dict_fuer_phase(phase, klasse, faecher_namen):
    faecher = {}
    for fach in faecher_namen:
        wt_dict = phase["lektionen"].get(klasse, {}).get(fach, {})
        faecher[fach] = [[wt, perioden] for wt, perioden in wt_dict.items()]
    return faecher


def erstelle_klassen_dateien(phasen, ferien, unterrichtsfreietage, spezialwochen, spezialtage, ausgabe_ordner):
    ausgabe_ordner.mkdir(parents=True, exist_ok=True)

    alle_klassen = sorted({klasse for p in phasen for klasse in p["lektionen"]})

    mitte = len(phasen) // 2
    semester = [("1. Semester", phasen[:mitte]), ("2. Semester", phasen[mitte:])]

    for klasse in alle_klassen:
        klassenstufe = klassenstufe_von_klasse(klasse)
        faecher_namen = sorted({fach for p in phasen for fach in p["lektionen"].get(klasse, {})})

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        if ist_konstant(phasen, klasse, faecher_namen):
            if len(faecher_namen) > 2:
                print(f"WARNUNG: Klasse {klasse} hat mehr als zwei Fächer {faecher_namen} - übersprungen.")
                continue
            faecher = faecher_dict_fuer_phase(phasen[0], klasse, faecher_namen)
            ws = wb.create_sheet("Jahresübersicht")
            fuelle_arbeitsblatt(ws, faecher, ferien, unterrichtsfreietage, spezialwochen, spezialtage, klassenstufe,
                                 phasen[0]["start"], phasen[-1]["ende"])
            beschreibung = "ganzes Jahr"
        else:
            uebersprungen = False
            erzeugte_blaetter = 0
            for semester_name, semester_phasen in semester:
                if ist_konstant(semester_phasen, klasse, faecher_namen):
                    if len(faecher_namen) > 2:
                        print(f"WARNUNG: Klasse {klasse} hat mehr als zwei Fächer {faecher_namen} - "
                              f"übersprungen.")
                        uebersprungen = True
                        break
                    faecher = faecher_dict_fuer_phase(semester_phasen[0], klasse, faecher_namen)
                    ws = wb.create_sheet(semester_name)
                    fuelle_arbeitsblatt(ws, faecher, ferien, unterrichtsfreietage, spezialwochen, spezialtage,
                                         klassenstufe, semester_phasen[0]["start"], semester_phasen[-1]["ende"])
                    erzeugte_blaetter += 1
                    continue

                for p in semester_phasen:
                    hat = any(eintraege for _, eintraege in klassen_signatur(p, klasse, faecher_namen))
                    if not hat:
                        continue  # Klasse wird in dieser Phase nicht unterrichtet - kein Arbeitsblatt
                    phase_faecher = sorted(
                        fach for fach in faecher_namen if p["lektionen"].get(klasse, {}).get(fach)
                    )
                    if len(phase_faecher) > 2:
                        print(f"WARNUNG: Klasse {klasse} hat in {p['name']} mehr als zwei Fächer "
                              f"{phase_faecher} - Klasse übersprungen.")
                        uebersprungen = True
                        break
                    faecher = faecher_dict_fuer_phase(p, klasse, phase_faecher)
                    ws = wb.create_sheet(p["name"])
                    fuelle_arbeitsblatt(ws, faecher, ferien, unterrichtsfreietage, spezialwochen, spezialtage,
                                         klassenstufe, p["start"], p["ende"], mit_inl=True)
                    erzeugte_blaetter += 1
                if uebersprungen:
                    break

            if uebersprungen or erzeugte_blaetter == 0:
                continue
            beschreibung = f"{erzeugte_blaetter} Blätter"

        pfad = ausgabe_ordner / f"Planung_{klasse}.xlsx"
        wb.save(pfad)
        print(f"Erstellt: {pfad} ({beschreibung}, Fächer: {', '.join(faecher_namen)})")


def erstelle_alle_planungen(output_dir, ausgabe_ordner):
    """Liest die Stundenplan-Exporte in output_dir und erzeugt daraus die
    Klassen-Planungsdateien in ausgabe_ordner. Schulkalender kommt aus
    config.py. Wird von main.py sowie vom eigenständigen Aufruf dieser
    Datei genutzt."""
    phasen = lade_phasen(output_dir, BASIS_JAHR)
    erstelle_klassen_dateien(phasen, FERIEN, UNTERRICHTSFREIETAGE, SPEZIALWOCHEN, SPEZIALTAGE, ausgabe_ordner)


if __name__ == "__main__":
    output_dir = Path(__file__).parent / "output_files"
    ausgabe_ordner = output_dir / "klassen"
    erstelle_alle_planungen(output_dir, ausgabe_ordner)
