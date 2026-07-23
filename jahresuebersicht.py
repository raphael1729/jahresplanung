"""
Erstellt eine klassenübergreifende Jahresübersicht (1 Excel-Datei):
  - Blatt "Jahresübersicht": zwei übereinander gestapelte Wochen-/Phasen-
    Raster (1. Semester = Phase 1-3, 2. Semester = Phase 4-6) mit einer
    leeren Platzhalterzelle pro Klasse, Fach und Phase zum manuellen
    Eintragen der Lehrplan-Themen (diese Themen stehen in keiner unserer
    Datenquellen). Ferien/unterrichtsfreie Tage/Spezialwochen sind wie in
    den Planungsdateien eingefärbt, Wochen ohne Unterricht für eine
    Klasse/Fach-Kombination grau.
  - Blatt "Statistik": gehaltene/ausgefallene Lektionen je Klasse, Fach und
    Phase (bzw. zusammengefasst pro Semester, wenn sich der Stundenplan
    innerhalb des Semesters nicht ändert - wie bei den Planungsdateien).
    Ferien zählen dabei nicht zu den Lektionen und damit auch nicht als
    Ausfall.

Farben/Schrift kommen aus config.STIL - dieselben wie in den
Planungsdateien (siehe rendering._baue_stile()). Die Phasen-Kopfzeile
verwendet dieselbe violette Farbe wie die Datumsspalten dort.

Öffentliche Schnittstelle für main.py: erstelle_jahresuebersicht_datei().
"""

from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from config import BASIS_JAHR, FERIEN, SPEZIALTAGE, SPEZIALWOCHEN, UNTERRICHTSFREIETAGE
from klassenplanung import ist_konstant, klassenstufe_von_klasse
from rendering import WOCHENTAG_MAP, _baue_stile, blockierte_zeitraeume, erweitertes_ende, stufe_und_titel
from stundenplan_parsing import lade_phasen

SPALTE_KLASSE = 1
SPALTE_FACH = 2
ERSTE_WOCHEN_SPALTE = 3
SCHRIFTGROESSE = 12

# Woche-Status -> welcher Fill-Schlüssel aus _baue_stile() verwendet wird.
# "leer" (Lücke zwischen zwei Phasen, kein Block) bekommt bewusst keine
# Füllung.
FILL_JE_STATUS = {
    "ferien": "fill_ferien",
    "frei": "fill_frei",
    "kein_unterricht": "fill_leer",
    "normal": "fill_lektion_1",
}


def _montag(tag):
    return tag - timedelta(days=tag.weekday())


def _wochen_spalten(block_start, block_ende):
    """dict: Montag-Datum -> Spaltenindex, für jede Woche zwischen
    block_start und block_ende (inkl. Lücken dazwischen, z. B. Ferien)."""
    erster, letzter = _montag(block_start), _montag(block_ende)
    spalten, montag, spalte = {}, erster, ERSTE_WOCHEN_SPALTE
    while montag <= letzter:
        spalten[montag] = spalte
        montag += timedelta(days=7)
        spalte += 1
    return spalten


def _phasen_spaltenbereich(phase, wochen_spalten):
    """(erste, letzte) Spalte, die die Wochen dieser Phase belegen."""
    start_montag, ende_montag = _montag(phase["start"]), _montag(phase["ende"])
    spalten = [s for m, s in wochen_spalten.items() if start_montag <= m <= ende_montag]
    return min(spalten), max(spalten)


def _spezialtage_relevant(spezialtage, klassenstufe):
    ergebnis = {}
    for name, zeitraum in spezialtage.items():
        stufe, titel = stufe_und_titel(name, alle_klassen_praefix=True)
        if stufe is not None and stufe != klassenstufe:
            continue
        ergebnis[titel] = zeitraum
    return ergebnis


def _spezialtag_status(tag, start_zeit, ende_zeit, spezialtage):
    """Wie periode_status() in rendering.py, aber mit dem reinen
    Spezialtag-Namen als Grund statt einer formatierten Zeitangabe."""
    p_start = datetime.combine(tag, start_zeit)
    p_ende = datetime.combine(tag, ende_zeit)
    for name, (sz_start, sz_ende) in spezialtage.items():
        if p_start >= sz_ende or p_ende <= sz_start:
            continue
        if sz_start <= p_start and sz_ende >= p_ende:
            return "ausfall_ganz", name
        return "ausfall_teil", name
    return "normal", None


def _block_info(tag, bloecke):
    """(typ, label) des Ferien-/Frei-/Spezialwochen-Blocks, der diesen Tag
    abdeckt, sonst None."""
    for s, e, typ, label in bloecke:
        if s <= tag <= e:
            return typ, label
    return None


def _block_endet_diese_woche(montag, bloecke):
    """(typ, label) eines Blocks, der VOR dieser Woche beginnt und INNERHALB
    dieser Woche endet, sonst None. Für Blöcke, die nicht Mo-Fr einer
    einzelnen Woche entsprechen (z. B. eine Testwoche Do-Mi statt Mo-Fr):
    die erste, nur teilweise betroffene Woche bleibt unmarkiert, die zweite
    (in der der Block endet) gilt als betroffen - auch wenn nicht alle
    Unterrichtstage dieser Woche im Block liegen."""
    for s, e, typ, label in bloecke:
        if s < montag <= e <= montag + timedelta(days=6):
            return typ, label
    return None


def _stile_mit_schriftgroesse(stile, groesse):
    """Baut aus den Stilen von rendering._baue_stile() eine Kopie mit fixer
    Schriftgrösse (Schriftart/-farbe/-schnitt bleiben gleich wie in der
    Planung) - die Jahresübersicht ist kompakter als die Planungsblätter
    und braucht daher eine eigene, kleinere Grösse."""
    neu = dict(stile)
    for schluessel in ("font_header", "font_datum", "font_normal", "font_frei"):
        alt = stile[schluessel]
        neu[schluessel] = Font(name=alt.name, size=groesse, bold=alt.bold, italic=alt.italic, color=alt.color)
    return neu


def _klassen_und_faecher(phasen):
    """Sortierte Liste von (klasse, [fach, ...]) über alle Phasen hinweg."""
    alle_klassen = sorted({klasse for p in phasen for klasse in p["lektionen"]})
    return [
        (klasse, sorted({fach for p in phasen for fach in p["lektionen"].get(klasse, {})}))
        for klasse in alle_klassen
    ]


def _wochen_status_zeile(phasen_teil, block_start, block_ende, klasse, fach, klassenstufe, ferien,
                          unterrichtsfreietage, spezialwochen, wochen_spalten, mit_inl):
    """dict: Montag-Datum -> (status, label). status ist einer von "ferien",
    "frei" (unterrichtsfrei/Spezialwoche), "kein_unterricht" (Klasse hat
    dieses Fach in der Phase nicht), "normal" (Platzhalter für Themen) oder
    "leer" (Woche gehört zu keiner Phase dieses Blocks)."""
    phase_je_woche = {}
    for phase in phasen_teil:
        for m in wochen_spalten:
            if _montag(phase["start"]) <= m <= _montag(phase["ende"]):
                phase_je_woche[m] = phase

    alle_wochentage = set()
    for phase in phasen_teil:
        alle_wochentage.update(phase["lektionen"].get(klasse, {}).get(fach, {}))

    # block_start/block_ende statt der eigenen Phasengrenzen, damit Blöcke in
    # der Lücke vor/nach diesem Semester (z. B. Skilager zwischen den
    # Semestern, NaWi-Woche nach der letzten Phase) mit gefunden werden.
    bloecke = blockierte_zeitraeume(ferien, unterrichtsfreietage, spezialwochen, block_start, block_ende,
                                     klassenstufe, alle_wochentage, mit_inl=mit_inl)

    # phase_kennung sorgt dafür, dass sich merge-fähige Wochen niemals über
    # eine Phasengrenze hinweg zusammenfassen, auch wenn Status+Grund
    # identisch sind (z. B. zwei direkt aufeinanderfolgende Phasen ohne
    # Ferien dazwischen) - jede Phase bekommt so ihre eigene Platzhalterzelle.
    status = {}
    for m in wochen_spalten:
        phase = phase_je_woche.get(m)
        phase_kennung = phase["name"] if phase is not None else None
        if phase is None:
            tage_dieser_woche = [m + timedelta(days=i) for i in range(5)]
        else:
            wochentag_perioden = phase["lektionen"].get(klasse, {}).get(fach, {})
            if not wochentag_perioden:
                status[m] = ("kein_unterricht", None, phase_kennung)
                continue
            tage_dieser_woche = [
                m + timedelta(days=i) for i in range(7)
                if WOCHENTAG_MAP[i] in wochentag_perioden and phase["start"] <= m + timedelta(days=i) <= phase["ende"]
            ]
            if not tage_dieser_woche:
                status[m] = ("kein_unterricht", None, phase_kennung)
                continue

        treffer = [_block_info(tag, bloecke) for tag in tage_dieser_woche]
        if all(treffer):
            typ, label = treffer[0]
            status[m] = ("ferien" if typ in ("ferien", "frei") else "frei", label, phase_kennung)
            continue

        block_am_ende = _block_endet_diese_woche(m, bloecke)
        if block_am_ende is not None:
            typ, label = block_am_ende
            status[m] = ("ferien" if typ in ("ferien", "frei") else "frei", label, phase_kennung)
        else:
            status[m] = ("normal", None, phase_kennung) if phase is not None else ("leer", None, phase_kennung)
    return status


def _schreibe_zeile_status(ws, zeile, wochen_spalten, status, stile, letzte_spalte):
    montage = sorted(wochen_spalten.items(), key=lambda kv: kv[1])
    i, n = 0, len(montage)
    while i < n:
        status_i, label_i, _phase_i = status[montage[i][0]]
        j = i
        while j + 1 < n and status[montage[j + 1][0]] == (status_i, label_i, _phase_i):
            j += 1
        spalte_i, spalte_j = montage[i][1], montage[j][1]
        if spalte_j > spalte_i:
            ws.merge_cells(start_row=zeile, start_column=spalte_i, end_row=zeile, end_column=spalte_j)
        fill_schluessel = FILL_JE_STATUS.get(status_i)
        for c in range(spalte_i, spalte_j + 1):
            z = ws.cell(row=zeile, column=c)
            z.border = stile["border_all"]
            if fill_schluessel:
                z.fill = stile[fill_schluessel]
        if status_i in ("ferien", "frei") and label_i:
            kopf = ws.cell(row=zeile, column=spalte_i, value=label_i)
            kopf.font = stile["font_frei"]
            # In der letzten Wochenspalte des Semesters kein Zeilenumbruch,
            # sonst wird nur diese eine Zeile höher als der Rest (z. B. bei
            # "NaWi-Woche"/"Sommersportlager" in einer schmalen Spalte).
            if spalte_j == letzte_spalte:
                kopf.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            else:
                kopf.alignment = stile["align_center"]
        i = j + 1


def _schreibe_block(ws, start_zeile, phasen_teil, block_start, block_ende, klassen_faecher, klassenstufen, ferien,
                     unterrichtsfreietage, spezialwochen, stile):
    wochen_spalten = _wochen_spalten(block_start, block_ende)
    letzte_spalte = max(wochen_spalten.values())
    phase_zeile, kopf_zeile = start_zeile, start_zeile + 1

    kopf_klasse = ws.cell(row=kopf_zeile, column=SPALTE_KLASSE, value="Klasse")
    kopf_fach = ws.cell(row=kopf_zeile, column=SPALTE_FACH, value="Fach")
    for zelle in (kopf_klasse, kopf_fach):
        zelle.font = stile["font_header"]
        zelle.fill = stile["fill_header"]
        zelle.alignment = stile["align_center"]
        zelle.border = stile["border_all"]

    for phase in phasen_teil:
        c0, c1 = _phasen_spaltenbereich(phase, wochen_spalten)
        if c1 > c0:
            ws.merge_cells(start_row=phase_zeile, start_column=c0, end_row=phase_zeile, end_column=c1)
        for c in range(c0, c1 + 1):
            zelle = ws.cell(row=phase_zeile, column=c)
            zelle.fill = stile["fill_datum"]  # violett, wie die Datumsspalten in der Planung
            zelle.border = stile["border_all"]
        titel_zelle = ws.cell(row=phase_zeile, column=c0, value=phase["name"])
        titel_zelle.font = stile["font_header"]
        titel_zelle.alignment = stile["align_center"]

    for montag, spalte in wochen_spalten.items():
        zelle = ws.cell(row=kopf_zeile, column=spalte, value=f"{montag.day}.{montag.month}.")
        zelle.font = stile["font_datum"]
        zelle.alignment = stile["align_center"]
        zelle.border = stile["border_all"]
        ws.column_dimensions[get_column_letter(spalte)].width = 13

    ws.column_dimensions[get_column_letter(SPALTE_KLASSE)].width = 13
    ws.column_dimensions[get_column_letter(SPALTE_FACH)].width = 13

    zeile = kopf_zeile + 1
    for klasse, faecher in klassen_faecher:
        start_zeile_klasse = zeile
        klassenstufe = klassenstufen[klasse]
        # Diese Klasse wird "in Phasen unterrichtet" (mit_inl), wenn ihr
        # Stundenplan (alle Fächer gemeinsam) innerhalb dieses Semesters
        # nicht konstant bleibt - dieselbe Prüfung wie in klassenplanung.py.
        mit_inl_klasse = not ist_konstant(phasen_teil, klasse, faecher)
        for fach in faecher:
            fach_zelle = ws.cell(row=zeile, column=SPALTE_FACH, value=fach)
            fach_zelle.font = stile["font_normal"]
            fach_zelle.alignment = stile["align_center"]
            fach_zelle.border = stile["border_all"]

            status = _wochen_status_zeile(phasen_teil, block_start, block_ende, klasse, fach, klassenstufe, ferien,
                                           unterrichtsfreietage, spezialwochen, wochen_spalten, mit_inl_klasse)
            _schreibe_zeile_status(ws, zeile, wochen_spalten, status, stile, letzte_spalte)
            zeile += 1

        if zeile - 1 > start_zeile_klasse:
            ws.merge_cells(start_row=start_zeile_klasse, start_column=SPALTE_KLASSE,
                            end_row=zeile - 1, end_column=SPALTE_KLASSE)
        klasse_zelle = ws.cell(row=start_zeile_klasse, column=SPALTE_KLASSE, value=klasse)
        klasse_zelle.font = stile["font_normal"]
        klasse_zelle.alignment = stile["align_center"]
        klasse_zelle.border = stile["border_all"]

    return zeile


def _schreibe_raster(ws, phasen, klassen_faecher, klassenstufen, ferien, unterrichtsfreietage, spezialwochen, stile):
    mitte = len(phasen) // 2
    # Dieselbe Erweiterung wie in klassenplanung.py: Semester 1 reicht bis
    # zum Tag vor Semester 2 (deckt z. B. ein Skilager in der Lücke
    # dazwischen ab), Semester 2 wird erweitert, falls kurz nach der
    # letzten Phase noch eine Ferien-/Spezialwoche beginnt (z. B. NaWi-
    # Woche nach Phase 6).
    letztes_ende = erweitertes_ende(phasen[-1]["ende"], ferien, unterrichtsfreietage, spezialwochen)
    bloecke = [
        (phasen[:mitte], phasen[0]["start"], phasen[mitte]["start"] - timedelta(days=1)),
        (phasen[mitte:], phasen[mitte]["start"], letztes_ende),
    ]

    zeile = 1
    for phasen_teil, block_start, block_ende in bloecke:
        zeile = _schreibe_block(ws, zeile, phasen_teil, block_start, block_ende, klassen_faecher, klassenstufen,
                                 ferien, unterrichtsfreietage, spezialwochen, stile)
        zeile += 2
    ws.freeze_panes = ws.cell(row=1, column=ERSTE_WOCHEN_SPALTE)


def lektionen_statistik(phase, klasse, fach, klassenstufe, ferien, unterrichtsfreietage, spezialwochen, spezialtage,
                         mit_inl):
    """
    Zählt für eine Klasse/Fach in einer Phase, wie viele Lektionen gehalten
    werden bzw. ganz oder teilweise ausfallen, inkl. Ausfallgründen. Ferien
    und Testwochen zählen dabei nicht zu den Lektionen und damit auch nicht
    als Ausfall. None, wenn die Klasse dieses Fach in dieser Phase gar nicht
    hat.
    """
    wochentag_perioden = phase["lektionen"].get(klasse, {}).get(fach, {})
    if not wochentag_perioden:
        return None

    spezialtage_relevant = _spezialtage_relevant(spezialtage, klassenstufe)
    bloecke = blockierte_zeitraeume(ferien, unterrichtsfreietage, spezialwochen, phase["start"], phase["ende"],
                                     klassenstufe, set(wochentag_perioden), mit_inl=mit_inl)

    gehalten = ausgefallen_ganz = ausgefallen_teil = 0
    gruende = {}

    tag = phase["start"]
    while tag <= phase["ende"]:
        perioden = wochentag_perioden.get(WOCHENTAG_MAP[tag.weekday()])
        if perioden:
            for start_zeit, ende_zeit in perioden:
                block = _block_info(tag, bloecke)
                if block and block[0] in ("ferien", "testwoche"):
                    continue  # Ferien/Testwochen zählen nicht zu den Lektionen, auch nicht als Ausfall
                if block:
                    typ, grund = "ausfall_ganz", block[1]
                else:
                    typ, grund = _spezialtag_status(tag, start_zeit, ende_zeit, spezialtage_relevant)

                if typ == "normal":
                    gehalten += 1
                elif typ == "ausfall_ganz":
                    ausgefallen_ganz += 1
                else:
                    ausgefallen_teil += 1
                if grund:
                    gruende[grund] = gruende.get(grund, 0) + 1
        tag += timedelta(days=1)

    total = gehalten + ausgefallen_ganz + ausgefallen_teil
    return {
        "total": total, "gehalten": gehalten,
        "ausgefallen_ganz": ausgefallen_ganz, "ausgefallen_teil": ausgefallen_teil,
        "gruende": gruende,
    }


def _kombiniere_statistik(phasen_teil, klasse, fach, klassenstufe, ferien, unterrichtsfreietage, spezialwochen,
                           spezialtage, mit_inl):
    """Summiert lektionen_statistik() über mehrere Phasen (z. B. ein
    Semester). None, wenn die Klasse das Fach in keiner dieser Phasen hat."""
    gesamt = {"total": 0, "gehalten": 0, "ausgefallen_ganz": 0, "ausgefallen_teil": 0, "gruende": {}}
    gefunden = False
    for phase in phasen_teil:
        stat = lektionen_statistik(phase, klasse, fach, klassenstufe, ferien, unterrichtsfreietage, spezialwochen,
                                    spezialtage, mit_inl)
        if stat is None:
            continue
        gefunden = True
        for schluessel in ("total", "gehalten", "ausgefallen_ganz", "ausgefallen_teil"):
            gesamt[schluessel] += stat[schluessel]
        for grund, anzahl in stat["gruende"].items():
            gesamt["gruende"][grund] = gesamt["gruende"].get(grund, 0) + anzahl
    return gesamt if gefunden else None


GRUENDE_SPALTE = 8


def _schreibe_statistik_zeile(ws, zeile, klasse, fach, phasen_name, stat, stile):
    gruende_text = ", ".join(f"{grund} ({anzahl})" for grund, anzahl in sorted(stat["gruende"].items()))
    werte = [klasse, fach, phasen_name, stat["total"], stat["gehalten"], stat["ausgefallen_ganz"],
             stat["ausgefallen_teil"], gruende_text or "-"]
    for c, wert in enumerate(werte, start=1):
        zelle = ws.cell(row=zeile, column=c, value=wert)
        zelle.font = stile["font_normal"]
        zelle.border = stile["border_all"]
        # Gründe nicht umbrechen, sonst wird nur diese eine Zeile höher als
        # der Rest, sobald ein Grund-Text die Spaltenbreite überschreitet.
        if c == GRUENDE_SPALTE:
            zelle.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        else:
            zelle.alignment = stile["align_center"]
    return len(werte[GRUENDE_SPALTE - 1])


def _fuelle_statistik_blatt(ws, phasen, klassen_faecher, klassenstufen, ferien, unterrichtsfreietage, spezialwochen,
                             spezialtage, stile):
    spalten_titel = ["Klasse", "Fach", "Phase", "Lektionen total", "Gehalten", "Ganz ausgefallen",
                      "Teilweise ausgefallen", "Gründe"]
    for c, text in enumerate(spalten_titel, start=1):
        zelle = ws.cell(row=1, column=c, value=text)
        zelle.font = stile["font_header"]
        zelle.fill = stile["fill_header"]
        zelle.alignment = stile["align_center"]
        zelle.border = stile["border_all"]

    mitte = len(phasen) // 2
    semester = [("1. Semester", phasen[:mitte]), ("2. Semester", phasen[mitte:])]

    zeile = 2
    laengste_gruende = len(spalten_titel[GRUENDE_SPALTE - 1])
    for klasse, faecher in klassen_faecher:
        klassenstufe = klassenstufen[klasse]
        for fach in faecher:
            for semester_name, semester_phasen in semester:
                # Wie in klassenplanung.py: "in Phasen unterrichtet" bezieht
                # sich auf die Klasse als Ganzes (alle Fächer), nicht nur
                # auf dieses eine Fach.
                mit_inl_klasse = not ist_konstant(semester_phasen, klasse, faecher)
                if ist_konstant(semester_phasen, klasse, [fach]):
                    stat = _kombiniere_statistik(semester_phasen, klasse, fach, klassenstufe, ferien,
                                                  unterrichtsfreietage, spezialwochen, spezialtage, mit_inl_klasse)
                    if stat is None or stat["total"] == 0:
                        continue
                    laenge = _schreibe_statistik_zeile(ws, zeile, klasse, fach, semester_name, stat, stile)
                    laengste_gruende = max(laengste_gruende, laenge)
                    zeile += 1
                else:
                    for phase in semester_phasen:
                        stat = lektionen_statistik(phase, klasse, fach, klassenstufe, ferien, unterrichtsfreietage,
                                                     spezialwochen, spezialtage, mit_inl_klasse)
                        if stat is None or stat["total"] == 0:
                            continue
                        laenge = _schreibe_statistik_zeile(ws, zeile, klasse, fach, phase["name"], stat, stile)
                        laengste_gruende = max(laengste_gruende, laenge)
                        zeile += 1

    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = 12
    # Gründe so breit wie der längste Eintrag, damit alles auf einer Zeile
    # Platz hat, statt einer fixen (mal zu schmalen, mal zu breiten) Breite.
    ws.column_dimensions[get_column_letter(GRUENDE_SPALTE)].width = laengste_gruende + 2
    ws.freeze_panes = "A2"


def erstelle_jahresuebersicht(phasen, ferien, unterrichtsfreietage, spezialwochen, spezialtage, ausgabe_pfad):
    klassen_faecher = _klassen_und_faecher(phasen)
    klassenstufen = {klasse: klassenstufe_von_klasse(klasse) for klasse, _ in klassen_faecher}
    stile = _stile_mit_schriftgroesse(_baue_stile(), SCHRIFTGROESSE)

    wb = openpyxl.Workbook()
    ws_raster = wb.active
    ws_raster.title = "Jahresübersicht"
    _schreibe_raster(ws_raster, phasen, klassen_faecher, klassenstufen, ferien, unterrichtsfreietage, spezialwochen,
                      stile)

    ws_statistik = wb.create_sheet("Statistik")
    _fuelle_statistik_blatt(ws_statistik, phasen, klassen_faecher, klassenstufen, ferien, unterrichtsfreietage,
                             spezialwochen, spezialtage, stile)

    ausgabe_pfad.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ausgabe_pfad)
    print(f"Erstellt: {ausgabe_pfad}")


def erstelle_jahresuebersicht_datei(output_dir, ausgabe_pfad):
    """Liest die Stundenplan-Exporte in output_dir und erzeugt daraus die
    Jahresübersicht. Schulkalender kommt aus config.py. Wird von main.py
    sowie vom eigenständigen Aufruf dieser Datei genutzt."""
    phasen = lade_phasen(output_dir, BASIS_JAHR)
    erstelle_jahresuebersicht(phasen, FERIEN, UNTERRICHTSFREIETAGE, SPEZIALWOCHEN, SPEZIALTAGE, ausgabe_pfad)


if __name__ == "__main__":
    output_dir = Path(__file__).parent / "output_files"
    erstelle_jahresuebersicht_datei(output_dir, output_dir / "Jahresuebersicht.xlsx")
