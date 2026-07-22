"""
Liest die Untis-Stundenplan-Exporte in output_files/*.xlsx (siehe
pdf_timetable_to_xlsx.py) und liefert pro Phase die Lektionen je Klasse,
Fach und Wochentag inkl. der realen Lektionszeiten (aus dem Zeitraster in
Zeile 7 der Exporte). InL- und Konf-Lektionen werden dabei ignoriert.

Öffentliche Schnittstelle für klassenplanung.py: lade_phasen().
"""

import re
from datetime import date, time

import openpyxl

ABK_WOCHENTAG = {
    "Mo": "Montag", "Di": "Dienstag", "Mi": "Mittwoch",
    "Do": "Donnerstag", "Fr": "Freitag", "Sa": "Samstag", "So": "Sonntag",
}

IGNORIERTE_FAECHER = {"InL", "KONF"}

PHASEN_CAPTION_RE = re.compile(
    r"Phase\s*(\d+)\s*\(\s*(\d{1,2})\.(\d{1,2})\.\s*-\s*(\d{1,2})\.(\d{1,2})\.\s*\)"
)

ZEIT_RE = re.compile(r"^(\d{1,2}):(\d{2})-(\d{1,2}):(\d{2})$")


def lektionsraster(ws, header_zeile=7):
    """dict: Spaltenindex -> (Start-Uhrzeit, End-Uhrzeit), gelesen aus der
    Kopfzeile mit den Lektionszeiten (z. B. '8:00-8:45')."""
    raster = {}
    for c in range(2, ws.max_column + 1):
        wert = ws.cell(row=header_zeile, column=c).value
        if not wert:
            continue
        m = ZEIT_RE.match(str(wert).strip())
        if not m:
            continue
        h1, mi1, h2, mi2 = (int(x) for x in m.groups())
        raster[c] = (time(h1, mi1), time(h2, mi2))
    return raster


def lektionen_pro_phase(ws):
    """
    Liest das Stundenplan-Raster (ab Zeile 8, Zeitraster aus Zeile 7) eines
    Phasen-Arbeitsblatts und liefert dict: klasse -> fach -> {wochentag:
    Tupel von (Start-Uhrzeit, End-Uhrzeit), eine je Lektion, chronologisch}.
    InL/Konf-Einträge und Zellen ohne Klasse+Fach (z. B. "KONF") werden
    ignoriert, ebenso Spalten ohne erkennbare Zeitangabe in Zeile 7.
    """
    raster = lektionsraster(ws)
    ergebnis = {}
    for r in range(8, ws.max_row + 1):
        wochentag = ABK_WOCHENTAG.get(ws.cell(row=r, column=1).value)
        if wochentag is None:
            continue

        for c in range(2, ws.max_column + 1):
            zeiten = raster.get(c)
            if zeiten is None:
                continue
            wert = ws.cell(row=r, column=c).value
            if not wert:
                continue
            teile = wert.split("\n")
            if len(teile) < 2:
                continue  # z. B. "KONF" ohne Klasse/Fach
            klasse, fach = teile[0], teile[1]
            if fach in IGNORIERTE_FAECHER or klasse in IGNORIERTE_FAECHER:
                continue
            perioden = ergebnis.setdefault(klasse, {}).setdefault(fach, {}).setdefault(wochentag, [])
            perioden.append(zeiten)

    for fach_dict in ergebnis.values():
        for wt_dict in fach_dict.values():
            for wochentag in wt_dict:
                wt_dict[wochentag] = tuple(sorted(wt_dict[wochentag]))
    return ergebnis


def lade_phasen(output_dir, basis_jahr):
    """
    Liest alle Stundenplan-Exporte 'Stundenplan_*.xlsx' in output_dir
    (siehe main.py), extrahiert pro Arbeitsblatt Phasenname, Datumsbereich
    (Jahreszahlen werden anhand von basis_jahr und der chronologischen
    Reihenfolge rekonstruiert) und die Lektionen. Liefert eine nach
    Phasennummer sortierte Liste von dicts.
    """
    roh = []
    for pfad in sorted(output_dir.glob("Stundenplan_*.xlsx")):
        wb = openpyxl.load_workbook(pfad, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            caption = ws.cell(row=5, column=1).value or ""
            m = PHASEN_CAPTION_RE.search(caption)
            if not m:
                continue
            nr, d1, m1, d2, m2 = (int(x) for x in m.groups())
            roh.append({
                "nr": nr, "d1": d1, "m1": m1, "d2": d2, "m2": m2,
                "lektionen": lektionen_pro_phase(ws),
            })
    roh.sort(key=lambda p: p["nr"])

    jahr = basis_jahr
    voriges = None
    phasen = []
    for p in roh:
        if voriges is not None and (p["m1"], p["d1"]) < voriges:
            jahr += 1
        start = date(jahr, p["m1"], p["d1"])
        voriges = (p["m1"], p["d1"])
        if (p["m2"], p["d2"]) < voriges:
            jahr += 1
        ende = date(jahr, p["m2"], p["d2"])
        voriges = (p["m2"], p["d2"])
        phasen.append({
            "nr": p["nr"], "name": f"Phase {p['nr']}",
            "start": start, "ende": ende, "lektionen": p["lektionen"],
        })
    return phasen
