"""
Baut das Excel-Layout einer Planungs-Datei: Kopfzeile, nach Kalenderwoche
ausgerichtete Lektionszeilen, Ferien-/Frei-/Spezialwochen-Zeiträume als
verschmolzene Einzeilen-Blöcke, optionale InL-Lektionen am Wochenanfang
sowie stunden-genaue Spezialtag-Ausfälle (ganz oder teilweise).

Öffentliche Schnittstelle für klassenplanung.py: fuelle_arbeitsblatt().
Gestaltungswerte (Schrift, Farben, Masse) kommen aus config.STIL.
"""

import re
from datetime import datetime, timedelta
from itertools import groupby
from typing import NamedTuple

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import STIL

WOCHENTAG_MAP = {
    0: "Montag", 1: "Dienstag", 2: "Mittwoch", 3: "Donnerstag",
    4: "Freitag", 5: "Samstag", 6: "Sonntag",
}

WOCHENTAG_ABK = {
    0: "Mo", 1: "Di", 2: "Mi", 3: "Do", 4: "Fr", 5: "Sa", 6: "So",
}

STUFE_UND_TITEL_RE = re.compile(r"^(\d+)\.\s*Klassen\s*(.+)$")
NUMMERIERUNG_RE = re.compile(r"^\d+\.\s*")
GB_PLUS_PRAEFIX = "GB-Plus Klassen "
GB_PLUS_STUFE = "gb_plus"


def stufe_und_titel(key, *, alle_klassen_praefix=False, strikt=False):
    """
    Liest aus einem Key wie '1. Klassen Regiowoche' die Klassenstufe (1) und
    den reinen Titel ohne Stufe ('Regiowoche').

    Ein 'GB-Plus Klassen '-Präfix wird immer erkannt (liefert Stufe
    GB_PLUS_STUFE) - gilt nicht für eine Klassenstufe, sondern für Klassen,
    die in Phasen unterrichtet werden (mit_inl=True beim Aufruf von
    fuelle_arbeitsblatt()). Eine führende Nummerierung im Titel (z. B. die
    '1.' in 'GB-Plus Klassen 1. Testwoche', nötig für eindeutige Keys im
    SPEZIALWOCHEN-Dict) wird dabei entfernt - angezeigt wird nur 'Testwoche'.
    alle_klassen_praefix=True erkennt zusätzlich ein 'Alle Klassen '-Präfix
    (liefert dann Stufe None, gilt für jede Klassenstufe) - genutzt für
    Spezialtage.
    strikt=True wirft einen ValueError, wenn der Key keinem der Formate
    entspricht, statt (None, key) zurückzugeben - genutzt für
    Spezialwochen, die immer eine Stufe haben müssen.
    """
    if key.startswith(GB_PLUS_PRAEFIX):
        titel = key[len(GB_PLUS_PRAEFIX):].strip()
        return GB_PLUS_STUFE, NUMMERIERUNG_RE.sub("", titel)
    if alle_klassen_praefix and key.startswith("Alle Klassen "):
        return None, key[len("Alle Klassen "):].strip()
    m = STUFE_UND_TITEL_RE.match(key)
    if m:
        return int(m.group(1)), m.group(2).strip()
    if strikt:
        raise ValueError(f"Key '{key}' hat nicht das Format '<Stufe>. Klassen <Titel>'.")
    return None, key


def periode_status(tag, start_zeit, ende_zeit, spezialtage):
    """
    Prüft eine einzelne Lektion (an einem konkreten Tag, mit Start-/Endzeit)
    gegen alle Spezialtage (Name -> (Start-Datetime, End-Datetime)).

    Liefert ein dict:
      {"typ": "normal"}
      {"typ": "ausfall_ganz", "label": <Spezialtag-Name>}
      {"typ": "ausfall_teil", "label": "<Zeitangaben Unterricht/Ausfall>"}

    Bei mehreren überlappenden Spezialtagen wird der erste Treffer verwendet.
    """
    p_start = datetime.combine(tag, start_zeit)
    p_ende = datetime.combine(tag, ende_zeit)

    for name, (sz_start, sz_ende) in spezialtage.items():
        if p_start >= sz_ende or p_ende <= sz_start:
            continue  # keine Überlappung

        if sz_start <= p_start and sz_ende >= p_ende:
            return {"typ": "ausfall_ganz", "label": name}

        grenz_start = max(p_start, sz_start)
        grenz_ende = min(p_ende, sz_ende)
        teile = []
        if p_start < grenz_start:
            teile.append(f"{p_start:%H:%M}-{grenz_start:%H:%M} Unterricht")
        teile.append(f"{grenz_start:%H:%M}-{grenz_ende:%H:%M} Ausfall: {name}")
        if grenz_ende < p_ende:
            teile.append(f"{grenz_ende:%H:%M}-{p_ende:%H:%M} Unterricht")
        return {"typ": "ausfall_teil", "label": ", ".join(teile)}

    return {"typ": "normal"}


def blockierte_zeitraeume(ferien, unterrichtsfreietage, spezialwochen, start, ende,
                           klassenstufe, wochentage_mit_unterricht, mit_inl=False):
    """
    Liefert eine nach Startdatum sortierte Liste (start, ende, typ, label) für
    alle Ferien-, unterrichtsfreien und (klassenstufen-relevanten) Spezial-
    Zeiträume, geschnitten auf [start, ende].

    Unterrichtsfreie Tage werden nur aufgenommen, wenn an diesem Wochentag
    für dieses Arbeitsblatt überhaupt Unterricht vorgesehen ist. Spezial-
    wochen werden nur aufgenommen, wenn ihre Klassenstufe zur Klasse passt -
    'GB-Plus Klassen'-Spezialwochen (siehe stufe_und_titel()) nur, wenn
    mit_inl=True (d. h. diese Klasse wird in Phasen unterrichtet).
    """
    bloecke = []
    for name, (s, e) in ferien.items():
        s2, e2 = max(s, start), min(e, ende)
        if s2 <= e2:
            bloecke.append((s2, e2, "ferien", name))
    for name, tag in unterrichtsfreietage.items():
        if start <= tag <= ende and WOCHENTAG_MAP[tag.weekday()] in wochentage_mit_unterricht:
            bloecke.append((tag, tag, "frei", name))
    for key, (s, e) in spezialwochen.items():
        stufe, titel = stufe_und_titel(key, strikt=True)
        if stufe == GB_PLUS_STUFE:
            if not mit_inl:
                continue
            typ = "testwoche"
        else:
            if stufe != klassenstufe:
                continue
            typ = "spezialwoche"
        s2, e2 = max(s, start), min(e, ende)
        if s2 <= e2:
            bloecke.append((s2, e2, typ, titel))
    bloecke.sort(key=lambda b: b[0])
    return bloecke


def ist_blockiert(tag, bloecke):
    return any(s <= tag <= e for s, e, _, _ in bloecke)


def erweitertes_ende(basis_ende, ferien, unterrichtsfreietage, spezialwochen, puffer_tage=14):
    """
    Erweitert ein Enddatum, falls kurz danach (innerhalb von puffer_tage)
    noch ein Ferien-/Frei-/Spezialwochen-Zeitraum beginnt (z. B. eine
    Abschlusswoche direkt nach der letzten Phase, wie NaWi-Woche/
    Sommersportlager/Maturaarbeitswoche nach Phase 6) - deckt ihn dann mit
    ab, statt ihn beim Rendering abzuschneiden. Prüft alle Klassenstufen
    zugleich (unabhängig von einer bestimmten Klasse), da es hier nur um
    die Render-Reichweite geht, nicht um deren tatsächliche Zuordnung.
    """
    erweitert = basis_ende
    for s, e in list(ferien.values()) + list(spezialwochen.values()):
        if basis_ende < s <= basis_ende + timedelta(days=puffer_tage):
            erweitert = max(erweitert, e)
    for tag in unterrichtsfreietage.values():
        if basis_ende < tag <= basis_ende + timedelta(days=puffer_tage):
            erweitert = max(erweitert, tag)
    return erweitert


def kalenderwoche(tag):
    """(ISO-Jahr, ISO-Kalenderwoche) – robust bei Jahreswechsel, nur intern für die Ausrichtung."""
    iso = tag.isocalendar()
    return (iso[0], iso[1])


def tage_pro_fach(eintraege, bloecke, spezialtage, start, ende):
    """
    Liefert chronologisch sortierte Liste von (datum, zeilen) für ein
    einzelnes Fach (eintraege = [[Wochentag, Perioden], ...], Perioden =
    Tupel von (Start-Uhrzeit, End-Uhrzeit) je Lektion), unter Auslassung
    aller Tage, die in einen blockierten Zeitraum (Ferien/frei) fallen.

    zeilen ist eine Liste (eine pro Lektion dieses Tages) von Status-dicts,
    siehe periode_status().
    """
    wochentag_perioden = dict(eintraege)

    tage = []
    tag = start
    while tag <= ende:
        wochentag = WOCHENTAG_MAP[tag.weekday()]
        if wochentag in wochentag_perioden and not ist_blockiert(tag, bloecke):
            zeilen = [periode_status(tag, s, e, spezialtage) for s, e in wochentag_perioden[wochentag]]
            tage.append((tag, zeilen))
        tag += timedelta(days=1)
    return tage


def gruppiere_nach_kw(tage):
    """dict: (ISO-Jahr, KW) -> Liste von Einträgen (nur zur internen Ausrichtung)."""
    return {kw: list(gruppe) for kw, gruppe in groupby(tage, key=lambda t: kalenderwoche(t[0]))}


def zeitspanne_kurz(start, ende):
    """Kompakte Anzeige ohne führende Nullen/Jahr, z. B. '28.9 - 11.10' oder '1.5'."""
    def kurz(tag):
        return f"{tag.day}.{tag.month}"

    if start == ende:
        return kurz(start)
    return f"{kurz(start)} - {kurz(ende)}"


class Spalten(NamedTuple):
    """Spaltenbelegung eines Arbeitsblatts: je Fach eine Datums- und eine
    (2-teilige) Datenspalte, siehe fuelle_arbeitsblatt()."""
    datum_1: int
    daten_1: int
    daten_2: int
    datum_2: int
    hat_fach2: bool


def platziere_fach_woche(ws, eintraege_woche, start_zeile, datum_spalte, daten_spalte, stile,
                          ziel_gesamtzeilen=None):
    """
    Platziert die Unterrichts-Einträge eines Fachs für EINE Kalenderwoche,
    beginnend bei start_zeile. Gibt die Anzahl belegter Zeilen zurück.

    eintraege_woche = [(datum, zeilen), ...], wobei zeilen eine Liste von
    Status-dicts ist (eines pro Lektion an diesem Tag, siehe periode_status()):
    normale Lektionen bleiben leer, ganz ausgefallene werden wie ein
    Frei-Tag markiert, teilweise ausgefallene bekommen eine eigene Farbe
    mit Zeitangabe. Mehrere direkt aufeinanderfolgende Lektionen mit
    demselben Ausfall-Status (gleicher Typ + gleiches Label) werden zu
    einer gemeinsamen, über alle betroffenen Zeilen verschmolzenen Zelle
    zusammengefasst.

    Ist an einem Tag JEDE Lektion vom selben Ausfall betroffen, bekommt auch
    die Datumszelle dessen Farbe (statt der neutralen Datums-Farbe) - bei
    gemischten Tagen (z. B. nur der Nachmittag fällt aus) bleibt sie neutral,
    da keine einzelne Farbe den Tag fair repräsentieren würde.

    ziel_gesamtzeilen erzwingt (falls gesetzt und grösser als die natürliche
    Lektionenzahl) eine Mindest-Gesamtzeilenzahl für die Woche: der
    Überschuss wird NICHT auf echte Lektionen verteilt (die bleiben alle
    gleich hoch), sondern als EINE leere, dezent gefüllte Platzhalterzelle
    ans Ende der Woche angehängt - genutzt, damit zwei Fächer mit
    unterschiedlich vielen Lektionen pro Woche optisch gleich viel Platz
    einnehmen, ohne dass eine einzelne Lektion künstlich gestreckt wirkt.
    """
    zeile = start_zeile
    for datum, zeilen in eintraege_woche:
        anzahl = len(zeilen)
        if anzahl > 1:
            ws.merge_cells(start_row=zeile, start_column=datum_spalte,
                            end_row=zeile + anzahl - 1, end_column=datum_spalte)
        datum_text = f"{WOCHENTAG_ABK[datum.weekday()]} {datum.day}.{datum.month}"
        datum_zelle = ws.cell(row=zeile, column=datum_spalte, value=datum_text)
        datum_zelle.alignment = stile["align_center"]
        datum_zelle.font = stile["font_datum"]

        if zeilen and all(z["typ"] != "normal" for z in zeilen) and all(z == zeilen[0] for z in zeilen):
            datum_fill = stile["fill_frei"] if zeilen[0]["typ"] == "ausfall_ganz" else stile["fill_teilausfall"]
        else:
            datum_fill = stile["fill_datum"]

        for i in range(anzahl):
            z = ws.cell(row=zeile + i, column=datum_spalte)
            z.border = stile["border_all"]
            z.fill = datum_fill

        i = 0
        while i < anzahl:
            info = zeilen[i]
            j = i
            while info["typ"] != "normal" and j + 1 < anzahl and zeilen[j + 1] == info:
                j += 1
            r0, r1 = zeile + i, zeile + j

            if info["typ"] == "normal":
                z2 = ws.cell(row=r0, column=daten_spalte)
                z3 = ws.cell(row=r0, column=daten_spalte + 1)
                for z in (z2, z3):
                    z.border = stile["border_all"]
                    z.font = stile["font_normal"]
                    z.fill = stile["fill_lektion"]
            else:
                fill = stile["fill_frei"] if info["typ"] == "ausfall_ganz" else stile["fill_teilausfall"]
                ws.merge_cells(start_row=r0, start_column=daten_spalte, end_row=r1, end_column=daten_spalte + 1)
                z2 = ws.cell(row=r0, column=daten_spalte, value=info["label"])
                z2.font = stile["font_frei"]
                z2.fill = fill
                z2.alignment = stile["align_center"]
                z2.border = stile["border_all"]
                for r in range(r0, r1 + 1):
                    for c in (daten_spalte, daten_spalte + 1):
                        z = ws.cell(row=r, column=c)
                        z.fill = fill
                        z.border = stile["border_all"]

            i = j + 1

        zeile += anzahl

    rest = max(0, (ziel_gesamtzeilen or 0) - (zeile - start_zeile))
    if rest > 0:
        ws.merge_cells(start_row=zeile, start_column=datum_spalte, end_row=zeile + rest - 1, end_column=datum_spalte)
        ws.merge_cells(start_row=zeile, start_column=daten_spalte, end_row=zeile + rest - 1, end_column=daten_spalte + 1)
        for r in range(zeile, zeile + rest):
            for c in (datum_spalte, daten_spalte, daten_spalte + 1):
                z = ws.cell(row=r, column=c)
                z.border = stile["border_all"]
                z.fill = stile["fill_leer"]
        zeile += rest

    return zeile - start_zeile


def platziere_block(ws, block, zeile, stile, spalten):
    """
    Platziert einen Ferien-/Frei-Zeitraum als EINE Zeile, über beide Fächer
    verschmolzen: Spalte A (und F, falls vorhanden) zeigen die Zeitspanne,
    die Spalten dazwischen (B:E bzw. B:C) werden zu einer Zelle mit dem Namen
    verschmolzen.
    """
    start, ende, typ, label = block
    fill = stile["fill_ferien"] if typ in ("ferien", "frei") else stile["fill_frei"]
    text_zeitspanne = zeitspanne_kurz(start, ende)

    datum_spalten = [spalten.datum_1, spalten.datum_2] if spalten.hat_fach2 else [spalten.datum_1]
    for spalte in datum_spalten:
        z = ws.cell(row=zeile, column=spalte, value=text_zeitspanne)
        z.font = stile["font_frei"]
        z.fill = fill
        z.alignment = stile["align_center"]
        z.border = stile["border_all"]

    ende_spalte = (spalten.daten_2 + 1) if spalten.hat_fach2 else (spalten.daten_1 + 1)
    ws.merge_cells(start_row=zeile, start_column=spalten.daten_1, end_row=zeile, end_column=ende_spalte)
    label_zelle = ws.cell(row=zeile, column=spalten.daten_1, value=label)
    label_zelle.font = stile["font_frei"]
    label_zelle.fill = fill
    label_zelle.alignment = stile["align_center"]
    label_zelle.border = stile["border_all"]
    for spalte in range(spalten.daten_1, ende_spalte + 1):
        z = ws.cell(row=zeile, column=spalte)
        z.fill = fill
        z.border = stile["border_all"]


def platziere_inl_block(ws, zeile, stile, spalten, anzahl=2):
    """
    Platziert 'anzahl' zusätzliche InL-Lektionen am Anfang einer
    Kalenderwoche: Statt eines Datums steht 'InL' in der Datumsspalte
    (über die Lektionszeilen verschmolzen), die Lektionszeilen selbst
    bleiben normale, leere Lektionszellen. Wird für beide Fächer (falls
    vorhanden) synchron auf denselben Zeilen platziert. Gibt die Anzahl
    belegter Zeilen zurück.
    """
    datum_spalten = [spalten.datum_1, spalten.datum_2] if spalten.hat_fach2 else [spalten.datum_1]
    for spalte in datum_spalten:
        if anzahl > 1:
            ws.merge_cells(start_row=zeile, start_column=spalte,
                            end_row=zeile + anzahl - 1, end_column=spalte)
        z = ws.cell(row=zeile, column=spalte, value="InL")
        z.alignment = stile["align_center"]
        z.font = stile["font_datum"]
        for i in range(anzahl):
            zi = ws.cell(row=zeile + i, column=spalte)
            zi.border = stile["border_all"]
            zi.fill = stile["fill_datum"]

    daten_spalten = [(spalten.daten_1, stile["fill_lektion_1"])]
    if spalten.hat_fach2:
        daten_spalten.append((spalten.daten_2, stile["fill_lektion_2"]))
    for daten_spalte, fill in daten_spalten:
        for i in range(anzahl):
            for c in (daten_spalte, daten_spalte + 1):
                z = ws.cell(row=zeile + i, column=c)
                z.border = stile["border_all"]
                z.font = stile["font_normal"]
                z.fill = fill

    return anzahl


def _baue_stile():
    """Baut alle für ein Arbeitsblatt benötigten openpyxl-Stile aus config.STIL."""
    def schrift(rolle):
        merkmale = STIL["schriften"].get(rolle, {})
        return Font(
            name=merkmale.get("font_name", STIL["font_name"]),
            size=merkmale.get("font_size", STIL["font_size"]),
            bold=merkmale.get("bold", False),
            italic=merkmale.get("italic", False),
            color=merkmale.get("farbe"),
        )

    return {
        "font_header": schrift("header"),
        "font_datum": schrift("datum"),
        "font_normal": schrift("normal"),
        "font_frei": schrift("frei"),
        "fill_lektion_1": PatternFill("solid", fgColor=STIL["farbe_daten"]),
        "fill_lektion_2": PatternFill("solid", fgColor=STIL["farbe_daten"]),
        "fill_datum": PatternFill("solid", fgColor=STIL["farbe_datum"]),
        "fill_ferien": PatternFill("solid", fgColor=STIL["farbe_ferien"]),
        "fill_frei": PatternFill("solid", fgColor=STIL["farbe_frei"]),
        "fill_teilausfall": PatternFill("solid", fgColor=STIL["farbe_teilausfall"]),
        "fill_leer": PatternFill("solid", fgColor=STIL["farbe_leer"]),
        "fill_header": PatternFill("solid", fgColor=STIL["farbe_header"]),
        "align_center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border_all": Border(*(Side(style="thin", color=STIL["farbe_rand"]),) * 4),
    }


def _schreibe_kopfzeile(ws, fach1, fach2, spalten, stile):
    """Schreibt Zeile 1: 'Datum' über den Datumsspalten, Fachnamen zentriert
    über ihren jeweiligen Datenspalten."""
    d1 = ws.cell(row=1, column=spalten.datum_1, value="Datum")
    d1.font = stile["font_header"]
    d1.fill = stile["fill_header"]
    d1.alignment = stile["align_center"]
    d1.border = stile["border_all"]

    ws.merge_cells(start_row=1, start_column=spalten.daten_1, end_row=1, end_column=spalten.daten_1 + 1)
    t1 = ws.cell(row=1, column=spalten.daten_1, value=fach1)
    t1.font = stile["font_header"]
    t1.fill = stile["fill_header"]
    t1.alignment = stile["align_center"]
    t1.border = stile["border_all"]
    z1b = ws.cell(row=1, column=spalten.daten_1 + 1)
    z1b.fill = stile["fill_header"]
    z1b.border = stile["border_all"]

    if fach2:
        ws.merge_cells(start_row=1, start_column=spalten.daten_2, end_row=1, end_column=spalten.daten_2 + 1)
        t2 = ws.cell(row=1, column=spalten.daten_2, value=fach2)
        t2.font = stile["font_header"]
        t2.fill = stile["fill_header"]
        t2.alignment = stile["align_center"]
        t2.border = stile["border_all"]
        z2b = ws.cell(row=1, column=spalten.daten_2 + 1)
        z2b.fill = stile["fill_header"]
        z2b.border = stile["border_all"]

        d2 = ws.cell(row=1, column=spalten.datum_2, value="Datum")
        d2.font = stile["font_header"]
        d2.fill = stile["fill_header"]
        d2.border = stile["border_all"]
        d2.alignment = stile["align_center"]


def _baue_zeitachse(alle_kw, kw_fach1, kw_fach2, bloecke):
    """
    Führt Unterrichtswochen und Ferien-/Frei-Blöcke chronologisch (nach
    Startdatum) zusammen. Liefert eine sortierte Liste von ("woche", kw)-
    bzw. ("block", block)-Tupeln.
    """
    ereignisse = []
    for kw in alle_kw:
        eintraege_woche = kw_fach1.get(kw, []) + kw_fach2.get(kw, [])
        erster_tag = min(datum for datum, _ in eintraege_woche)
        ereignisse.append((erster_tag, "woche", kw))
    for block in bloecke:
        ereignisse.append((block[0], "block", block))
    ereignisse.sort(key=lambda e: e[0])
    return [(art, daten) for _, art, daten in ereignisse]


def fuelle_arbeitsblatt(ws, faecher, ferien, unterrichtsfreietage, spezialwochen, spezialtage,
                         klassenstufe, start, ende, mit_inl=False):
    """
    Füllt ein (bereits vorhandenes) Arbeitsblatt: Kopfzeile mit Fachnamen,
    danach Datenzeilen nach Kalenderwoche ausgerichtet, Ferien-/Frei-/
    Spezialwochen-Zeiträume als verschmolzene Einzeilen-Blöcke. Spezialtage
    (Name -> [Start-Datetime, End-Datetime]) blocken keinen ganzen Tag,
    sondern werden pro Lektion gegen deren reale Uhrzeit geprüft (siehe
    periode_status()).

    mit_inl=True fügt an den Anfang jeder Kalenderwoche 2 zusätzliche
    InL-Lektionen ein (statt eines Datums steht 'InL', siehe
    platziere_inl_block()) - gedacht für Klassen, die in mehrere Phasen
    unterteilt sind.
    """
    fach_liste = list(faecher.keys())
    if len(fach_liste) > 2:
        raise ValueError("Es sind maximal zwei Fächer erlaubt.")

    fach1 = fach_liste[0]
    fach2 = fach_liste[1] if len(fach_liste) > 1 else None

    wochentage_mit_unterricht = {wt for eintraege in faecher.values() for wt, _ in eintraege}

    bloecke = blockierte_zeitraeume(ferien, unterrichtsfreietage, spezialwochen, start, ende,
                                     klassenstufe, wochentage_mit_unterricht, mit_inl=mit_inl)

    spezialtage_relevant = {}
    for name, zeitraum in spezialtage.items():
        stufe, titel = stufe_und_titel(name, alle_klassen_praefix=True)
        if stufe is not None and stufe != klassenstufe:
            continue
        spezialtage_relevant[titel] = zeitraum

    tage_fach1 = tage_pro_fach(faecher[fach1], bloecke, spezialtage_relevant, start, ende)
    tage_fach2 = tage_pro_fach(faecher[fach2], bloecke, spezialtage_relevant, start, ende) if fach2 else []

    kw_fach1 = gruppiere_nach_kw(tage_fach1)
    kw_fach2 = gruppiere_nach_kw(tage_fach2)
    alle_kw = sorted(set(kw_fach1) | set(kw_fach2))

    stile = _baue_stile()
    spalten = Spalten(datum_1=1, daten_1=2, daten_2=4, datum_2=6, hat_fach2=bool(fach2))

    _schreibe_kopfzeile(ws, fach1, fach2, spalten, stile)

    stile_fach1 = dict(stile, fill_lektion=stile["fill_lektion_1"])
    stile_fach2 = dict(stile, fill_lektion=stile["fill_lektion_2"])

    zeile = 2
    for art, daten in _baue_zeitachse(alle_kw, kw_fach1, kw_fach2, bloecke):
        if art == "woche":
            if mit_inl:
                zeile += platziere_inl_block(ws, zeile, stile, spalten)
            kw = daten
            eintraege1 = kw_fach1.get(kw, [])
            eintraege2 = kw_fach2.get(kw, []) if fach2 else []
            ziel_gesamtzeilen = max(sum(len(z) for _, z in eintraege1), sum(len(z) for _, z in eintraege2))
            rows1 = platziere_fach_woche(ws, eintraege1, zeile, spalten.datum_1, spalten.daten_1, stile_fach1,
                                          ziel_gesamtzeilen=ziel_gesamtzeilen)
            rows2 = 0
            if fach2:
                rows2 = platziere_fach_woche(ws, eintraege2, zeile, spalten.datum_2, spalten.daten_2, stile_fach2,
                                              ziel_gesamtzeilen=ziel_gesamtzeilen)
            zeile += max(rows1, rows2, 1)
        else:
            platziere_block(ws, daten, zeile, stile, spalten)
            zeile += 1

    # --- Spaltenbreiten setzen  --------------------------------------------------
    breite_datum, breite_daten = STIL["spaltenbreite_datum"], STIL["spaltenbreite_daten"]
    ws.column_dimensions[get_column_letter(spalten.datum_1)].width = breite_datum
    ws.column_dimensions[get_column_letter(spalten.daten_1)].width = breite_daten
    ws.column_dimensions[get_column_letter(spalten.daten_1 + 1)].width = breite_daten
    if fach2:
        ws.column_dimensions[get_column_letter(spalten.daten_2)].width = breite_daten
        ws.column_dimensions[get_column_letter(spalten.daten_2 + 1)].width = breite_daten
        ws.column_dimensions[get_column_letter(spalten.datum_2)].width = breite_datum

    # --- Zeilenhöhen setzen -------------------------------------------------------
    ws.row_dimensions[1].height = STIL["zeilenhoehe_header"]
    for r in range(2, zeile):
        ws.row_dimensions[r].height = STIL["zeilenhoehe"]

    # --- Trennlinie zwischen den beiden Fächern verdicken --------------------------
    if fach2:
        dick = Side(style="thick", color=STIL["farbe_rand"])
        spalte_links, spalte_rechts = spalten.daten_1 + 1, spalten.daten_2
        for r in range(1, zeile):
            for spalte, kante in ((spalte_links, "right"), (spalte_rechts, "left")):
                zelle = ws.cell(row=r, column=spalte)
                rand = zelle.border
                kanten = {"left": rand.left, "right": rand.right, "top": rand.top, "bottom": rand.bottom}
                kanten[kante] = dick
                zelle.border = Border(**kanten)
