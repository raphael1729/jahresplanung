#!/usr/bin/env python3
"""
Convert an Untis-style school timetable PDF (one ruled grid per page/phase)
into an Excel workbook with one sheet per page, laid out just like the PDF.

Usage:
    python pdf_timetable_to_xlsx.py input.pdf [output.xlsx]

Requires: pdfplumber, openpyxl  (pip install pdfplumber openpyxl)
"""

import re
import sys
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def cluster_lines(words, tol=3):
    """Group words into text lines based on their vertical position."""
    lines = []
    for w in sorted(words, key=lambda w: (w["top"], w["x0"])):
        if lines and abs(w["top"] - lines[-1][0]["top"]) <= tol:
            lines[-1].append(w)
        else:
            lines.append([w])
    return [" ".join(x["text"] for x in sorted(line, key=lambda w: w["x0"])) for line in lines]


def extract_header_footer(page, table_bbox):
    """Grab the free-text title block above the table and the caption below it."""
    _, top, _, bottom = table_bbox
    words = page.extract_words()
    above = [w for w in words if w["top"] < top - 1]
    below = [w for w in words if w["top"] > bottom + 1]

    mid_x = page.width / 2
    left_lines = cluster_lines([w for w in above if w["x0"] < mid_x])
    right_lines = cluster_lines([w for w in above if w["x0"] >= mid_x])
    footer_lines = cluster_lines(below)

    return left_lines, right_lines, footer_lines


def parse_page(page):
    """Return (grid, left_header_lines, footer_caption) for one page of the PDF."""
    tables = page.find_tables()
    if not tables:
        return None
    table = tables[0]
    grid = table.extract()
    left_lines, _right_lines, footer_lines = extract_header_footer(page, table.bbox)
    caption = footer_lines[0] if footer_lines else ""
    return grid, left_lines, caption


def sanitize_sheet_name(name, used):
    name = re.sub(r'[\[\]:*?/\\]', "", name).strip() or "Sheet"
    name = name[:31]
    base, i = name, 1
    while name in used:
        suffix = f" ({i})"
        name = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(name)
    return name


def build_workbook(pdf_path):
    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            parsed = parse_page(page)
            if not parsed:
                continue
            grid, left_lines, caption = parsed

            n_cols = max(len(row) for row in grid)
            n_rows = len(grid)

            sheet_title = re.sub(r"\s*\(.*\)\s*$", "", caption).strip() or f"Page {len(wb.sheetnames) + 1}"
            ws = wb.create_sheet(sanitize_sheet_name(sheet_title, used_names))

            # Free-text header block, reproduced as-is above the grid.
            row_cursor = 1
            for line in left_lines:
                ws.cell(row=row_cursor, column=1, value=line).font = Font(name="Arial", size=10, bold=(row_cursor == 1))
                row_cursor += 1
            if caption:
                ws.cell(row=row_cursor, column=1, value=caption).font = Font(name="Arial", size=10, italic=True)
                row_cursor += 1
            row_cursor += 1  # blank spacer row

            header_row = row_cursor
            for col_idx, value in enumerate(grid[0], start=1):
                c = ws.cell(row=header_row, column=col_idx, value=value or None)
                c.font = Font(name="Arial", size=9, bold=True)
                c.alignment = CENTER
                c.fill = HEADER_FILL
                c.border = BORDER

            for r, row in enumerate(grid[1:], start=1):
                out_row = header_row + r
                for col_idx, value in enumerate(row, start=1):
                    c = ws.cell(row=out_row, column=col_idx, value=value or None)
                    c.border = BORDER
                    c.alignment = CENTER
                    if col_idx == 1:
                        c.font = Font(name="Arial", size=11, bold=True)
                        c.fill = HEADER_FILL
                    else:
                        c.font = Font(name="Arial", size=9)
                ws.row_dimensions[out_row].height = 55

            ws.column_dimensions["A"].width = 8
            for col_idx in range(2, n_cols + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 12
            ws.row_dimensions[header_row].height = 20
            ws.sheet_view.showGridLines = False

    return wb


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    pdf_path = Path(sys.argv[1])
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else pdf_path.with_suffix(".xlsx")

    wb = build_workbook(pdf_path)
    wb.save(out_path)
    print(f"Saved {out_path}")


if __name__ == "__main__":
    main()
